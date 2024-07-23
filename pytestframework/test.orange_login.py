#task-12
import pytest
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from selenium.webdriver.chrome.service import Service as ChromeService
import sys
import os
import time

# Ensure the parent directory is in the system path for module imports
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from orange_login import OrangeLoginPage


@pytest.fixture
def browser():
    chromedriver_path = r"C:\Users\ajay\OneDrive\Desktop\chromedriver.exe"
    driver = webdriver.Chrome(service=chrome_service)
    driver.maximize_window()
    yield driver
    time.sleep(6)
    driver.quit()


def read_test_data_from_excel(file_path=r"C:\Users\ajayk\OneDrive\Documents\Book1.xlsx", sheet_name='Sheet1'):
    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    test_data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[2] is not None:  # Ensure both username and password are present
            test_data.append((row[1], row[2]))
        else:
            break

    return test_data, wb, ws



def test_orange_login(browser, username, password):
    login_page = OrangeLoginPage(browser)

    try:
        # Navigate to the orange login page
        login_page.navigate_to_login_page()

        # Wait for the sign-in button to be visible
        WebDriverWait(browser, 10).until(EC.visibility_of_element_located(login_page.sign_in_button))
        #
        login_page.click_login_button()


        # Wait for username input field to be visible
        WebDriverWait(browser, 10).until(EC.visibility_of_element_located(login_page.username_locator))

        login_page.enter_username(username)
        login_page.continue_username()

        # Wait for password input field to be visible
        WebDriverWait(browser, 10).until(EC.visibility_of_element_located(login_page.password_locator))

        login_page.enter_password(password)
        login_page.click_login_button()

        # Wait for login success (customize as per your application)
        WebDriverWait(browser, 10).until(EC.title_contains("orange HRM"))
        return True
    except Exception as e:
        print(f"Error during login process: {e}")
        return False
        raise


@pytest.mark.parametrize("username,password", read_test_data_from_excel()[0])
def test_orange_login_data_driven(browser, username, password):
    test_data,wb,ws = read_test_data_from_excel()
    for row in ws.iter_rows(min_row=2):
        if row[1].value==username and row[2].value==password:
            row_index = row[1].row
            break
    if test_orange_login(browser, username, password):
        ws.cell(row=row_index,column=7).value="login successful"
    else:
        ws.cell(row=row_index,column=7).value="login failed"
    wb.save(r"c:\Users\ajayk\OneDrive\Documents\Book1.xlsx")



if __name__ == "__main__":
    pytest.main(["-v", "test_orange_login.py"])